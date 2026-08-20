"""Retailaufgaben – Qualifikations-/Ampelmatrix fürs Team.

Zeilen = Aufgaben, Spalten = Teammitglieder, Zellen = Ampel-Selbsteinschätzung:
  grün  = Aufgabe kann ich
  gelb  = brauche noch Hilfe/Unterstützung
  rot   = kann ich nicht / noch nie gemacht

Ziel: die Teamleitung sieht, wo für Urlaubs-/Krankheitsvertretung noch
Schulungsbedarf ist (Aufgaben, die niemand oder nur eine Person „grün" kann).

Für alle angemeldeten Rollen (Team = Retail, Teamleitung = Admin). Team
identifiziert sich per Namensauswahl (kein Einzel-Login).
"""

import io
import json
import os
import secrets
from datetime import datetime
from pathlib import Path

from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.responses import (FileResponse, HTMLResponse, JSONResponse,
                               RedirectResponse, Response)
from starlette.background import BackgroundTask
from starlette.middleware.base import BaseHTTPMiddleware

import auth

APP_NAME = 'Retailaufgaben'
ACCENT = '#7c3aed'
TYP = 'RA-13'

DATA_DIR = Path(os.environ.get('RA_DATA_DIR', '/data'))
AUFGABEN_PATH = DATA_DIR / 'aufgaben.json'
PERSONEN_PATH = DATA_DIR / 'personen.json'
BEWERTUNG_PATH = DATA_DIR / 'bewertung.json'
HISTORIE_PATH = DATA_DIR / 'historie.json'
NOTIZEN_PATH = DATA_DIR / 'notizen.json'        # {'tid|pid': [{ts, text}]}
MASSNAHMEN_PATH = DATA_DIR / 'massnahmen.json'  # {'tid|pid': {termin, pate, status, erstellt}}
AUDIT_DB = DATA_DIR / 'audit.db'

# Ampelstufen: Schlüssel -> (Farbe, kurz, lang)
AMPEL = {
    'gruen': ('#16a34a', 'Grün', 'kann ich'),
    'gelb': ('#eab308', 'Gelb', 'brauche Hilfe'),
    'rot': ('#dc2626', 'Rot', 'kann ich nicht'),
}
STUFEN = ('gruen', 'gelb', 'rot')
# Rang für „gelernt/Rückschritt": höher = mehr Kompetenz
RANG = {'': 0, 'rot': 1, 'gelb': 2, 'gruen': 3}

app = FastAPI(title=APP_NAME, docs_url=None, redoc_url=None)


# ── Persistenz ───────────────────────────────────────────────────────────────
def _lade(pfad, standard):
    try:
        return json.loads(pfad.read_text(encoding='utf-8'))
    except Exception:  # noqa: BLE001
        return standard


def _sichere(pfad, inhalt):
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        pfad.write_text(json.dumps(inhalt, ensure_ascii=False), encoding='utf-8')
    except OSError:
        pass


def _aufgaben():
    liste = _lade(AUFGABEN_PATH, [])
    return liste if isinstance(liste, list) else []


def _personen(nur_aktiv=False):
    liste = _lade(PERSONEN_PATH, [])
    liste = liste if isinstance(liste, list) else []
    if nur_aktiv:
        return [p for p in liste if p.get('aktiv', True)]
    return liste  # Datei-Reihenfolge = Anzeige-Reihenfolge (per Hoch/Runter änderbar)


def _ist_backup(p) -> bool:
    return bool(p.get('backup'))


def _heute() -> str:
    return datetime.now().date().isoformat()


def _abgelaufen(p) -> bool:
    """Person nur befristet im Team - Stichtag vorbei?

    Eine Praktikantin, die im Oktober geht, darf im November nicht mehr als
    Abdeckung dastehen. Ohne Stichtag muesste jemand daran denken; mit
    Stichtag faellt sie von selbst heraus.
    """
    bis = str(p.get('bis') or '').strip()
    return bool(bis) and bis < _heute()


def _zaehlt_nicht_grund(p) -> str:
    if not p.get('aktiv', True):
        return 'inaktiv'
    if _abgelaufen(p):
        return f'nur bis {_datum_de(p.get("bis"))}'
    if _ist_backup(p):
        return (p.get('hinweis') or '').strip() or 'Backup'
    return ''


def _datum_de(iso) -> str:
    t = str(iso or '')
    return f'{t[8:10]}.{t[5:7]}.{t[0:4]}' if len(t) >= 10 else t


def _zaehlende(personen=None):
    """Wer zählt für die Abdeckung?

    Nicht mitgezählt werden: inaktive, befristete nach ihrem Stichtag und
    alle, die als „zählt nicht" markiert sind. Joern springt nur ein, wenn
    alle Stricke reissen; Mareike ist Praktikantin auf Zeit. Zaehlten sie mit,
    saehe jede Aufgabe, die nur sie koennen, wie abgedeckt aus - und genau die
    ist das Risiko. Sie bleiben sichtbar, fliessen aber nicht in
    „kritisch / Risiko / ok" ein.
    """
    liste = personen if personen is not None else _personen()
    return [p for p in liste
            if p.get('aktiv', True) and not _ist_backup(p) and not _abgelaufen(p)]


def _is_admin(user):
    return bool(user) and (user.get('role') == 'admin')


def _bewertungen():
    d = _lade(BEWERTUNG_PATH, {})
    return d if isinstance(d, dict) else {}


def _historie():
    h = _lade(HISTORIE_PATH, [])
    return h if isinstance(h, list) else []


def _historie_add(events):
    """events: Liste von {ts, pid, tid, alt, neu}. Nur echte Änderungen."""
    events = [e for e in events if e.get('alt') != e.get('neu')]
    if not events:
        return
    h = _historie()
    h.extend(events)
    if len(h) > 3000:
        h = h[-3000:]
    _sichere(HISTORIE_PATH, h)


def _notizen():
    d = _lade(NOTIZEN_PATH, {})
    return d if isinstance(d, dict) else {}


def _massnahmen():
    d = _lade(MASSNAHMEN_PATH, {})
    return d if isinstance(d, dict) else {}


def _ckey(tid, pid):
    return f'{tid}|{pid}'


def _tage_her(ts):
    """Ganze Tage seit einem 'YYYY-MM-DD ...'-Zeitstempel (oder None)."""
    try:
        d = datetime.strptime(str(ts)[:10], '%Y-%m-%d')
        return (datetime.now() - d).days
    except Exception:  # noqa: BLE001
        return None


def _esc(s):
    return (str(s if s is not None else '').replace('&', '&amp;')
            .replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;'))


# ── Zugang (alle angemeldeten Rollen) ────────────────────────────────────────
async def ra_middleware(request, call_next):
    user = auth.get_current_user(request)
    request.state.user = user
    if request.url.path in ('/login', '/health'):
        return await call_next(request)
    if user:
        return await call_next(request)
    if request.method == 'GET':
        return RedirectResponse('/login', status_code=303)
    return JSONResponse(status_code=401, content={'detail': 'Nicht angemeldet.'})


app.add_middleware(BaseHTTPMiddleware, dispatch=ra_middleware)


@app.get('/login', response_class=HTMLResponse)
def login_page():
    return HTMLResponse(auth.login_page_html(app_name=APP_NAME))


@app.post('/login')
async def login_submit(request: Request):
    form = await request.form()
    user = auth.authenticate(form.get('email', ''), form.get('password', ''))
    if not user:
        return HTMLResponse(auth.login_page_html(error='E-Mail oder Passwort falsch', app_name=APP_NAME))
    token = auth.create_token(user['email'], user['role'])
    resp = RedirectResponse('/', status_code=303)
    resp.set_cookie(auth.COOKIE_NAME, token, max_age=auth.SESSION_HOURS * 3600,
                    httponly=True, samesite='lax')
    return resp


@app.get('/logout')
def logout(request: Request):
    resp = RedirectResponse('/login', status_code=303)
    resp.delete_cookie(auth.COOKIE_NAME)
    return resp


@app.get('/health')
def health():
    return {'status': 'ok'}


# ── Seiten-Shell ─────────────────────────────────────────────────────────────
def _seite(inhalt, user=None):
    kopf = auth.user_header_html(user) if user else ''
    return (
        '<!DOCTYPE html><html lang="de"><head><meta charset="utf-8">'
        '<meta name="viewport" content="width=device-width, initial-scale=1">'
        f'<title>{APP_NAME} &middot; Skyport Werkbank</title>'
        '<link rel="stylesheet" href="https://skyport-werkbank.sliplane.app/werkbank.css">'
        f'<style>:root{{--accent:{ACCENT}}}'
        '.subtabs{display:flex;gap:2px;border-bottom:2px solid var(--line);margin:20px 0 4px;flex-wrap:wrap}'
        '.subtabs a{padding:11px 20px;font-family:var(--cond);font-weight:600;text-transform:uppercase;'
        'letter-spacing:.08em;font-size:12.5px;color:var(--muted);text-decoration:none;'
        'border-bottom:2px solid transparent;margin-bottom:-2px}'
        '.subtabs a.on{color:var(--accent);border-bottom-color:var(--accent)}'
        '.subtabs a:hover{color:var(--ink)}'
        '.mtx{width:100%;border-collapse:collapse;font-size:13px}'
        '.mtx th,.mtx td{border:1px solid var(--line);padding:7px 9px;text-align:center;vertical-align:middle}'
        '.mtx th{background:#f4f2fb;position:sticky;top:0}'
        '.mtx td.auf{text-align:left;min-width:240px;font-weight:400;padding-left:26px;position:relative}'
        '.mtx td.auf::before{content:"–";position:absolute;left:11px;color:var(--muted)}'
        '.mtx .ber{background:#f0ecfa;text-align:left;font-weight:700;font-size:13.5px;color:var(--ink)}'
        '.dot{display:inline-block;width:15px;height:15px;border-radius:50%;border:1px solid rgba(0,0,0,.12)}'
        '.dot.none{background:#eef1f4}'
        '.chip2{display:inline-block;padding:1px 8px;border-radius:20px;font-size:11px;font-weight:600}'
        '.arow{display:flex;align-items:center;gap:12px;padding:9px 0;border-bottom:1px solid var(--line);flex-wrap:wrap}'
        '.arow .an{flex:1;min-width:200px;font-size:14px}'
        '.arow.tp .an{padding-left:18px;position:relative}'
        '.arow.tp .an::before{content:"–";position:absolute;left:2px;color:var(--muted)}'
        '.grp{font-weight:700;font-size:14.5px;color:var(--ink);margin:18px 0 2px;padding-top:8px;'
        'border-top:1px solid var(--line)}.grp:first-child{border-top:0;margin-top:2px}'
        '.ampel{display:flex;gap:6px;flex-wrap:wrap}'
        '.ampel input{width:auto;height:auto;margin:0}'
        '.pill{display:inline-flex;align-items:center;gap:6px;padding:6px 11px;border:1px solid var(--line);'
        'border-radius:20px;cursor:pointer;font-size:12.5px;user-select:none;background:#fff}'
        '.pill:has(input:checked){color:#fff;border-color:transparent}'
        '.pill.g:has(input:checked){background:#16a34a}.pill.y:has(input:checked){background:#eab308}'
        '.pill.r:has(input:checked){background:#dc2626}.pill.n:has(input:checked){background:#6b7280}'
        '.tscroll{overflow-x:auto}'
        '</style></head><body>'
        '<header class="bar"><div class="brand">'
        '<a href="https://skyport-werkbank.sliplane.app">Skyport <b>&middot;</b> Werkbank</a></div>'
        '<nav>'
        '<a href="https://skyport-werkbank.sliplane.app">&Uuml;bersicht</a>'
        '<a class="on" href="/">Retailaufgaben</a>'
        '</nav>'
        f'<div class="live">{kopf}</div></header>'
        '<div class="band"></div><main>' + inhalt + '</main></body></html>')


def _platte(sub=''):
    return ('<div class="plate"><div><div class="eyebrow">Werkzeug</div>'
            f'<h1>{APP_NAME}</h1>'
            + (f'<p class="sub">{sub}</p>' if sub else '') +
            f'</div><div class="typ"><span>Typ</span><strong>{TYP}</strong></div></div>')


def _subtabs(active, is_admin=False):
    def a(key, label, href):
        cls = ' class="on"' if active == key else ''
        return f'<a href="{href}"{cls}>{label}</a>'
    tabs = (a('matrix', 'Übersicht', '/')
            + a('bewerten', 'Meine Bewertung', '/bewerten')
            + a('verwalten', 'Aufgaben & Team', '/verwalten'))
    if is_admin:
        tabs += a('verlauf', 'Verlauf & Lernbedarf', '/verlauf')
    return '<div class="subtabs">' + tabs + '</div>'


def _dot(stufe):
    if stufe in AMPEL:
        return f'<span class="dot" style="background:{AMPEL[stufe][0]}" title="{AMPEL[stufe][2]}"></span>'
    return '<span class="dot none" title="keine Angabe"></span>'


def _legende():
    teile = ' &nbsp; '.join(f'{_dot(s)} {AMPEL[s][1]} &ndash; {AMPEL[s][2]}' for s in STUFEN)
    return f'<p class="hint" style="margin:0 0 8px">{teile} &nbsp; {_dot("")} keine Angabe</p>'


# ── Übersicht / Matrix ───────────────────────────────────────────────────────
def _matrix_html(is_admin=False):
    aufgaben = _aufgaben()
    personen = _personen(nur_aktiv=True)
    bew = _bewertungen()

    if not aufgaben:
        return (_platte('Ampelmatrix: wer kann welche Aufgabe – für Urlaubs- und Krankheitsvertretung')
                + _subtabs('matrix', is_admin)
                + '<div class="card"><p class="hint" style="margin:0">Noch keine Aufgaben. '
                'Lege sie unter <a href="/verwalten">Aufgaben &amp; Team</a> an oder importiere deine '
                'Excel-Tabelle.</p></div>')

    massn = _massnahmen()
    notz = _notizen()

    def stufe(tid, pid):
        return (bew.get(str(tid)) or {}).get(str(pid), '')

    def zelle(tid, pid):
        s = stufe(tid, pid)
        ck = _ckey(tid, pid)
        mark = ''
        m = massn.get(ck)
        if m and m.get('status') != 'erledigt' and (m.get('termin') or m.get('pate')):
            mark += '<span title="Maßnahme geplant" style="color:#7c3aed">&#9679;</span>'
        if notz.get(ck):
            mark += '<span title="Notiz vorhanden" style="color:#6b7280">&#9998;</span>'
        inner = _dot(s) + (f'<span style="margin-left:3px;font-size:10px">{mark}</span>' if mark else '')
        if is_admin:
            return (f'<td><a href="/zelle/{_esc(tid)}/{_esc(pid)}" '
                    f'style="text-decoration:none;display:inline-block">{inner}</a></td>')
        return f'<td>{inner}</td>'

    # Schulungsbedarf: Aufgaben mit < 2 „grün" sind Vertretungsrisiko.
    # Gezählt wird ohne Backup-Personen - siehe _zaehlende().
    zaehlt = _zaehlende(personen)
    krit, risk = [], []
    for a in aufgaben:
        g = sum(1 for p in zaehlt if stufe(a['id'], p['id']) == 'gruen')
        if g == 0:
            krit.append(a)
        elif g == 1:
            risk.append(a)

    warn = (
        '<div class="row" style="margin:0 0 14px;gap:10px">'
        f'<span class="chip2" style="background:#fde8e8;color:#a10e0e">{len(krit)} kritisch '
        '(niemand „grün")</span>'
        f'<span class="chip2" style="background:#fdf3d6;color:#8a5a00">{len(risk)} Risiko '
        '(nur 1 Person „grün")</span>'
        + (f'<span class="hint" style="margin:0">Gez&auml;hlt werden '
           + ', '.join(_esc(x['name']) for x in zaehlt)
           + '. Nicht gez&auml;hlt: '
           + ', '.join(f'{_esc(x["name"])} ({_esc(_zaehlt_nicht_grund(x))})'
                       for x in personen if _zaehlt_nicht_grund(x))
           + '.</span>'
           if any(_zaehlt_nicht_grund(x) for x in personen) else '')
        + f'<span class="chip2" style="background:#e7f6ec;color:#0f7a3d">'
        f'{len(aufgaben) - len(krit) - len(risk)} abgedeckt (≥ 2)</span></div>')

    if not personen:
        koerper = ('<div class="card"><p class="hint" style="margin:0">Noch keine Teammitglieder. '
                   'Unter <a href="/verwalten">Aufgaben &amp; Team</a> anlegen, dann kann jeder unter '
                   '„Meine Bewertung" seine Ampel setzen.</p></div>')
    else:
        kopf = '<th class="auf" style="text-align:left">Aufgabe</th>'
        for p in personen:
            grund = _zaehlt_nicht_grund(p)
            if grund:
                kopf += (f'<th style="opacity:.7">{_esc(p["name"])}'
                         '<div style="font-size:9px;font-weight:600;letter-spacing:.04em;'
                         f'color:#6d28d9">{_esc(grund.upper())}</div></th>')
            else:
                kopf += f'<th>{_esc(p["name"])}</th>'
        kopf += '<th>grün</th><th>Status</th>'

        # nach Bereich gruppieren (Reihenfolge der ersten Vorkommen)
        bereiche, gesehen = [], set()
        for a in aufgaben:
            b = (a.get('bereich') or '').strip()
            if b not in gesehen:
                gesehen.add(b)
                bereiche.append(b)

        n_sp = len(personen) + 3
        zeilen = ''
        for b in bereiche:
            if b:
                zeilen += f'<tr><td class="ber" colspan="{n_sp}">{_esc(b)}</td></tr>'
            for a in [x for x in aufgaben if (x.get('bereich') or '').strip() == b]:
                g = sum(1 for p in zaehlt if stufe(a['id'], p['id']) == 'gruen')
                if g == 0:
                    stat = '<span class="chip2" style="background:#fde8e8;color:#a10e0e">kritisch</span>'
                elif g == 1:
                    stat = '<span class="chip2" style="background:#fdf3d6;color:#8a5a00">Risiko</span>'
                else:
                    stat = '<span class="chip2" style="background:#e7f6ec;color:#0f7a3d">ok</span>'
                zellen = ''.join(zelle(a['id'], p['id']) for p in personen)
                zeilen += (f'<tr><td class="auf">{_esc(a["name"])}</td>{zellen}'
                           f'<td>{g}</td><td>{stat}</td></tr>')
        klick_hint = ('<p class="hint" style="margin:10px 0 0">Tipp: Klick auf einen Ampelpunkt öffnet '
                      'Notizen &amp; Maßnahme (Termin, Pate) für diese Aufgabe/Person. '
                      '<span style="color:#7c3aed">&#9679;</span> = Maßnahme geplant, '
                      '<span style="color:#6b7280">&#9998;</span> = Notiz.</p>') if is_admin else ''
        koerper = (f'<div class="card"><div class="tscroll"><table class="mtx"><thead><tr>{kopf}</tr>'
                   f'</thead><tbody>{zeilen}</tbody></table></div>'
                   '<div class="row" style="margin-top:12px">'
                   '<a class="btn secondary" href="/export.xlsx">Als Excel exportieren</a></div>'
                   f'{klick_hint}</div>')

    return (_platte('Ampelmatrix: wer kann welche Aufgabe – für Urlaubs- und Krankheitsvertretung')
            + _subtabs('matrix', is_admin) + _legende() + warn + koerper)


# ── Meine Bewertung ──────────────────────────────────────────────────────────
def _bewerten_html(pid='', is_admin=False):
    aufgaben = _aufgaben()
    personen = _personen(nur_aktiv=True)
    bew = _bewertungen()
    pid = str(pid or '')

    opt = '<option value="">— bitte wählen —</option>' + ''.join(
        f'<option value="{p["id"]}"{" selected" if str(p["id"]) == pid else ""}>{_esc(p["name"])}</option>'
        for p in personen)
    auswahl = (
        '<div class="card"><label class="fld"><span>Ich bin</span>'
        '<select onchange="location.href=\'/bewerten?person=\'+this.value" '
        'style="padding:9px 11px;border:1px solid #d5dde5;border-radius:7px;font:inherit;max-width:320px">'
        f'{opt}</select></label>'
        '<p class="hint" style="margin:6px 0 0">Wähle deinen Namen und setze bei jeder Aufgabe deine '
        'Ampel. Grün = kann ich, Gelb = brauche Hilfe, Rot = kann ich nicht / noch nie gemacht.</p></div>')

    if not personen:
        return (_platte('Meine Selbsteinschätzung') + _subtabs('bewerten', is_admin)
                + '<div class="card"><p class="hint" style="margin:0">Noch keine Teammitglieder – '
                'zuerst unter <a href="/verwalten">Aufgaben &amp; Team</a> anlegen.</p></div>')
    if not pid or not any(str(p['id']) == pid for p in personen):
        return (_platte('Meine Selbsteinschätzung') + _subtabs('bewerten', is_admin) + auswahl)
    if not aufgaben:
        return (_platte('Meine Selbsteinschätzung') + _subtabs('bewerten', is_admin) + auswahl
                + '<div class="card"><p class="hint" style="margin:0">Noch keine Aufgaben angelegt.</p></div>')

    # nach Hauptaufgabe gruppieren (Reihenfolge des ersten Vorkommens)
    bereiche, gesehen = [], set()
    for a in aufgaben:
        b = (a.get('bereich') or '').strip()
        if b not in gesehen:
            gesehen.add(b)
            bereiche.append(b)

    rows = ''
    for b in bereiche:
        if b:
            rows += f'<div class="grp">{_esc(b)}</div>'
        for a in [x for x in aufgaben if (x.get('bereich') or '').strip() == b]:
            cur = (bew.get(str(a['id'])) or {}).get(pid, '')

            def radio(val, cls, label, _a=a, _cur=cur):
                ch = ' checked' if _cur == val else ''
                return (f'<label class="pill {cls}"><input type="radio" name="t{_a["id"]}" '
                        f'value="{val}"{ch}>{label}</label>')

            rows += (
                '<div class="arow tp">'
                f'<div class="an">{_esc(a["name"])}</div>'
                '<div class="ampel">'
                + radio('gruen', 'g', 'kann ich')
                + radio('gelb', 'y', 'brauche Hilfe')
                + radio('rot', 'r', 'kann ich nicht')
                + radio('', 'n', '—')
                + '</div></div>')

    schnell = (
        '<form method="post" action="/bewerten/pauschal" '
        'onsubmit="return confirm(\'Alle Aufgaben dieser Person auf einmal überschreiben?\')" '
        'class="card" style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-bottom:14px">'
        f'<input type="hidden" name="person" value="{_esc(pid)}">'
        '<b style="font-size:13px">Alle auf einmal:</b>'
        '<button class="btn secondary btn-sm" name="stufe" value="gruen" type="submit">alle grün</button>'
        '<button class="btn secondary btn-sm" name="stufe" value="gelb" type="submit">alle gelb</button>'
        '<button class="btn secondary btn-sm" name="stufe" value="rot" type="submit">alle rot</button>'
        '<button class="btn secondary btn-sm" name="stufe" value="" type="submit">zurücksetzen</button>'
        '<span class="hint" style="margin:0">Setzt jede Teilaufgabe – danach unten einzeln feinjustieren.</span>'
        '</form>')
    form = (
        '<form method="post" action="/bewerten">'
        f'<input type="hidden" name="person" value="{_esc(pid)}">'
        f'<div class="card">{rows}'
        '<div class="row" style="margin-top:14px"><button class="btn" type="submit">Speichern</button> '
        '<span class="hint" style="align-self:center">Deine Angaben, jederzeit änderbar.</span></div>'
        '</div></form>')
    return (_platte('Meine Selbsteinschätzung') + _subtabs('bewerten', is_admin) + auswahl + schnell + form)


# ── Verwalten (Aufgaben, Team, Import/Export) ────────────────────────────────
def _verwalten_html(meldung='', is_admin=False):
    aufgaben = _aufgaben()
    personen = _personen()

    bereiche_a, gesehen_a = [], set()
    for a in aufgaben:
        b = (a.get('bereich') or '').strip()
        if b not in gesehen_a:
            gesehen_a.add(b)
            bereiche_a.append(b)
    a_zeilen = ''
    for b in bereiche_a:
        if b:
            a_zeilen += f'<div class="grp">{_esc(b)}</div>'
        for a in [x for x in aufgaben if (x.get('bereich') or '').strip() == b]:
            a_zeilen += (
                '<div class="row" style="justify-content:space-between;align-items:center;'
                'border-bottom:1px solid var(--line);padding:7px 0 7px 18px;margin:0">'
                f'<span>&ndash; {_esc(a["name"])}</span>'
                f'<form method="post" action="/aufgabe/{_esc(a["id"])}/loeschen" '
                'onsubmit="return confirm(\'Aufgabe löschen? Auch alle Bewertungen dazu.\')">'
                '<button class="btn secondary btn-sm" type="submit">löschen</button></form></div>')
        if b:
            a_zeilen += (
                '<form method="post" action="/aufgabe" class="row" style="gap:6px;'
                'padding:5px 0 12px 18px;margin:0;border-bottom:1px solid var(--line)">'
                f'<input type="hidden" name="bereich" value="{_esc(b)}">'
                '<input name="name" placeholder="+ Unteraufgabe hinzufügen" required '
                'style="flex:1;min-width:150px;padding:7px 10px;border:1px solid #d5dde5;'
                'border-radius:7px;font:inherit">'
                '<button class="btn secondary btn-sm" type="submit">hinzufügen</button></form>')

    n = len(personen)
    p_zeilen = ''
    for i, p in enumerate(personen):
        aktiv = p.get('aktiv', True)
        backup = _ist_backup(p)
        grund = _zaehlt_nicht_grund(p)
        chip = ('' if aktiv else
                ' <span class="chip2" style="background:#eef1f4;color:#6b7280">inaktiv</span>')
        if grund and aktiv:
            chip += (' <span class="chip2" style="background:#f3e8ff;color:#6d28d9" '
                     f'title="z&auml;hlt nicht in die Abdeckung">{_esc(grund)}</span>')
        if p.get('bis') and not _abgelaufen(p):
            chip += (' <span class="chip2" style="background:#eef1fe;color:#3730a3">bis '
                     f'{_esc(_datum_de(p.get("bis")))}</span>')
        name_html = (f'<span style="{"" if aktiv else "opacity:.55"}">{_esc(p["name"])}</span>{chip}')
        if is_admin:
            pid = _esc(p['id'])
            up_dis = ' disabled' if i == 0 else ''
            dn_dis = ' disabled' if i == n - 1 else ''
            ctrl = (
                '<div class="row" style="gap:5px;margin:0">'
                f'<form method="post" action="/person/{pid}/hoch" style="display:inline;margin:0">'
                f'<button class="btn secondary btn-sm" type="submit" title="nach oben"{up_dis}>&uarr;</button></form>'
                f'<form method="post" action="/person/{pid}/runter" style="display:inline;margin:0">'
                f'<button class="btn secondary btn-sm" type="submit" title="nach unten"{dn_dis}>&darr;</button></form>'
                f'<form method="post" action="/person/{pid}/aktiv" style="display:inline;margin:0">'
                f'<button class="btn secondary btn-sm" type="submit">{"deaktivieren" if aktiv else "aktivieren"}'
                '</button></form>'
                f'<form method="post" action="/person/{pid}/backup" style="display:inline;margin:0">'
                f'<button class="btn secondary btn-sm" type="submit" '
                'title="zählt dann nicht in die Abdeckung">'
                f'{"z&auml;hlt wieder mit" if backup else "z&auml;hlt nicht"}</button></form>'
                f'<form method="post" action="/person/{pid}/befristung" '
                'style="display:inline-flex;gap:4px;margin:0;align-items:center">'
                f'<input name="hinweis" value="{_esc(p.get("hinweis") or "")}" '
                'placeholder="Grund" style="width:120px;padding:5px 7px;'
                'border:1px solid #d5dde5;border-radius:6px;font:inherit;font-size:12px">'
                f'<input name="bis" type="date" value="{_esc(p.get("bis") or "")}" '
                'style="padding:5px 7px;border:1px solid #d5dde5;border-radius:6px;'
                'font:inherit;font-size:12px" title="nur bis zu diesem Tag im Team">'
                '<button class="btn secondary btn-sm" type="submit">merken</button></form>'
                f'<form method="post" action="/person/{pid}/loeschen" style="display:inline;margin:0" '
                'onsubmit="return confirm(\'Person löschen? Auch alle Bewertungen dieser Person.\')">'
                '<button class="btn secondary btn-sm" type="submit">löschen</button></form></div>')
        else:
            ctrl = ''
        p_zeilen += (
            '<div class="row" style="justify-content:space-between;align-items:center;gap:10px;'
            f'border-bottom:1px solid var(--line);padding:7px 0;margin:0"><span>{name_html}</span>{ctrl}</div>')

    a_block = a_zeilen or '<p class="hint">Noch keine Aufgaben.</p>'
    p_block = p_zeilen or '<p class="hint">Noch keine Personen.</p>'

    # Bestehende Hauptaufgaben (= Bereiche) als Vorschlagsliste zum Zuordnen
    haupt, gesehen = [], set()
    for a in aufgaben:
        b = (a.get('bereich') or '').strip()
        if b and b.lower() not in gesehen:
            gesehen.add(b.lower())
            haupt.append(b)
    haupt_opts = ''.join(f'<option value="{_esc(b)}">' for b in haupt)

    if is_admin:
        team_form = (
            '<form method="post" action="/person" class="row" style="gap:8px;flex-wrap:wrap">'
            '<input name="name" placeholder="Name" required '
            'style="flex:1;min-width:200px;padding:9px 11px;border:1px solid #d5dde5;border-radius:7px;font:inherit">'
            '<button class="btn" type="submit">+ Person</button></form>')
        team_hint = ('<p class="hint" style="margin:8px 0 0">Mit &uarr;&darr; sortieren, &bdquo;deaktivieren&ldquo; '
                     'nimmt jemanden aus Matrix &amp; Bewertung (Daten bleiben erhalten).</p>')
        import_card = (
            '<div class="sect"><h2>Excel</h2><div class="rule"></div></div>'
            '<div class="card"><p class="hint" style="margin-top:0">Import: Excel/CSV mit einer Spalte '
            '<b>Teilaufgabe</b> (erkannt werden auch &bdquo;Name&ldquo;, &bdquo;Aufgabe&ldquo;, '
            '&bdquo;Tätigkeit&ldquo;, &bdquo;Task&ldquo;); optionale Spalte <b>Hauptaufgabe</b> '
            '(&bdquo;Section&ldquo;, &bdquo;Column&ldquo;, &bdquo;Bereich&ldquo;, &bdquo;Kategorie&ldquo;) und '
            'optionale Spalte <b>Assignee</b> (&bdquo;Bearbeiter&ldquo;, &bdquo;Verantwortlich&ldquo;). '
            'Genannte Assignees werden automatisch als Person angelegt und ihre Aufgaben direkt auf '
            '<b>grün</b> gesetzt (jederzeit änderbar). Vorhandene Aufgaben bleiben erhalten, Doppelte '
            'werden übersprungen.</p>'
            '<form method="post" action="/import" enctype="multipart/form-data" class="row" style="gap:8px">'
            '<input type="file" name="datei" accept=".xlsx,.xlsm,.csv" required>'
            '<button class="btn" type="submit">Aufgaben importieren</button></form>'
            '<div class="row" style="margin-top:10px"><a class="btn secondary" href="/export.xlsx">'
            'Matrix als Excel exportieren</a></div></div>')
    else:
        team_form = ''
        team_hint = '<p class="hint" style="margin:8px 0 0">Das Team wird von der Teamleitung gepflegt.</p>'
        import_card = (
            '<div class="sect"><h2>Excel</h2><div class="rule"></div></div>'
            '<div class="card"><div class="row" style="margin:0"><a class="btn secondary" '
            'href="/export.xlsx">Matrix als Excel exportieren</a></div></div>')

    return (
        _platte('Aufgaben und Team pflegen')
        + _subtabs('verwalten', is_admin)
        + meldung
        + '<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(300px,1fr));gap:16px">'
        # Aufgaben (Hauptaufgabe bündelt Teilaufgaben)
        '<div class="card"><h2 style="margin-top:0">Aufgaben</h2>'
        '<form method="post" action="/aufgabe" class="row" style="gap:8px;flex-wrap:wrap">'
        '<input name="name" placeholder="Teilaufgabe (z. B. Wayfair)" required '
        'style="flex:1;min-width:200px;padding:9px 11px;border:1px solid #d5dde5;border-radius:7px;font:inherit">'
        '<input name="bereich" list="haupt" placeholder="Hauptaufgabe (z. B. Portale täglich)" '
        'style="width:210px;padding:9px 11px;border:1px solid #d5dde5;border-radius:7px;font:inherit">'
        f'<datalist id="haupt">{haupt_opts}</datalist>'
        '<button class="btn" type="submit">+ Aufgabe</button></form>'
        '<p class="hint" style="margin:8px 0 0">Gleiche <b>Hauptaufgabe</b> eintippen (oder aus der Liste '
        'wählen), um mehrere <b>Teilaufgaben</b> darunter zu bündeln – z.&nbsp;B. „Portale täglich" mit '
        'Pharao, Wayfair, Viking … Das Wie steht in Asana, hier zählt nur die Ampel-Einschätzung.</p>'
        f'<div style="margin-top:10px">{a_block}</div></div>'
        # Team
        '<div class="card"><h2 style="margin-top:0">Team</h2>'
        + team_form
        + f'<div style="margin-top:10px">{p_block}</div>'
        + team_hint
        + '</div>'
        '</div>'
        # Import / Export
        + import_card)


# ── Verlauf & Lernbedarf (nur Teamleitung/Admin) ─────────────────────────────
def _verlauf_html(f=''):
    aufgaben = _aufgaben()
    personen = _personen()
    aktive = _zaehlende(personen)
    bew = _bewertungen()
    hist = _historie()
    pname = {str(p['id']): p.get('name', '?') for p in personen}
    tinfo = {str(a['id']): (a.get('name') or '', a.get('bereich') or '') for a in aufgaben}
    n = len(aufgaben)
    massn = _massnahmen()

    def stufe(tid, pid):
        return (bew.get(str(tid)) or {}).get(str(pid), '')

    # ── Wochen-Report (letzte 7 Tage) + aktuelle Abdeckung ──
    gelernt7 = rueck7 = 0
    for e in hist:
        th = _tage_her(e.get('ts'))
        if th is not None and th <= 7:
            d = RANG.get(e.get('neu', ''), 0) - RANG.get(e.get('alt', ''), 0)
            if d > 0:
                gelernt7 += 1
            elif d < 0:
                rueck7 += 1
    krit = risk = 0
    for a in aufgaben:
        g = sum(1 for p in aktive if stufe(a['id'], p['id']) == 'gruen')
        if g == 0:
            krit += 1
        elif g == 1:
            risk += 1
    wochen_card = (
        '<div class="card" style="display:flex;gap:10px;flex-wrap:wrap;align-items:center">'
        f'<span class="chip2" style="background:#e7f6ec;color:#0f7a3d">{gelernt7}&times; gelernt (7 Tage)</span>'
        f'<span class="chip2" style="background:#fdecec;color:#b42318">{rueck7}&times; Rückschritt (7 Tage)</span>'
        '<span style="color:var(--line)">|</span>'
        f'<span class="chip2" style="background:#fde8e8;color:#a10e0e">{krit} kritisch</span>'
        f'<span class="chip2" style="background:#fdf3d6;color:#8a5a00">{risk} Risiko</span>'
        f'<span class="chip2" style="background:#e7f6ec;color:#0f7a3d">{len(aufgaben) - krit - risk} abgedeckt (≥ 2)</span>'
        '</div>')

    # ── Vertretungs-Warnung: Aufgaben mit < 2 grün, längste ohne Bewegung zuerst ──
    risiko = []
    for a in aufgaben:
        tid = str(a['id'])
        g = sum(1 for p in aktive if stufe(tid, p['id']) == 'gruen')
        if g >= 2:
            continue
        gelb = sum(1 for p in aktive if stufe(tid, p['id']) == 'gelb')
        tage = [_tage_her(e.get('ts')) for e in hist if str(e.get('tid')) == tid]
        tage = [x for x in tage if x is not None]
        letzte = min(tage) if tage else None  # kleinste Tageszahl = jüngste Bewegung
        risiko.append((a, g, gelb, letzte))
    risiko.sort(key=lambda r: (r[1], -(r[3] if r[3] is not None else 9999)))
    if risiko:
        rrows = ''
        for a, g, gelb, letzte in risiko:
            stat = ('<span class="chip2" style="background:#fde8e8;color:#a10e0e">kritisch</span>'
                    if g == 0 else '<span class="chip2" style="background:#fdf3d6;color:#8a5a00">Risiko</span>')
            if letzte is None:
                beweg = 'noch keine Bewegung'
            elif letzte >= 14:
                beweg = f'seit {letzte} Tg. ohne Bewegung'
            else:
                beweg = f'zuletzt vor {letzte} Tg.'
            rrows += (
                '<div class="arow" style="gap:10px">'
                f'<span style="flex:1;min-width:200px"><span style="color:var(--muted);font-size:11px">'
                f'{_esc(a.get("bereich") or "")}</span> {_esc(a.get("name") or "")}</span>'
                f'<span style="min-width:160px">{g}&times; grün · {gelb}&times; gelb (lernt)</span>'
                f'<span style="min-width:150px;color:var(--muted);font-size:12px">{beweg}</span>{stat}</div>')
        risiko_card = f'<div class="card">{rrows}</div>'
    else:
        risiko_card = ('<div class="card"><p class="hint" style="margin:0">Alle Aufgaben sind von '
                       '&ge; 2 Personen abgedeckt – kein Vertretungsrisiko.</p></div>')

    # ── Offene Maßnahmen (nach Termin) ──
    heute = datetime.now().strftime('%Y-%m-%d')
    offene = []
    for ck, mv in massn.items():
        if mv.get('status') == 'erledigt' or not (mv.get('termin') or mv.get('pate')):
            continue
        tidk, _sep, pidk = ck.partition('|')
        offene.append((tidk, pidk, mv))
    offene.sort(key=lambda x: x[2].get('termin') or '9999-99-99')
    if offene:
        mrows = ''
        for tidk, pidk, mv in offene:
            nm, ber = tinfo.get(str(tidk), ('gelöschte Aufgabe', ''))
            wer = pname.get(str(pidk), '?')
            pate = pname.get(str(mv.get('pate')), '') if mv.get('pate') else ''
            termin = mv.get('termin') or ''
            if termin and termin < heute:
                tchip = f'<span class="chip2" style="background:#fdecec;color:#b42318">überfällig {_esc(termin)}</span>'
            elif termin:
                tchip = f'<span class="chip2" style="background:#eef1f4;color:#374151">Termin {_esc(termin)}</span>'
            else:
                tchip = ''
            patetxt = f' · Pate: {_esc(pate)}' if pate else ''
            mrows += (
                '<div class="arow" style="gap:10px">'
                f'<a href="/zelle/{_esc(tidk)}/{_esc(pidk)}" style="flex:1;min-width:200px;text-decoration:none">'
                f'<span style="color:var(--muted);font-size:11px">{_esc(ber)}</span> {_esc(nm)}</a>'
                f'<span style="min-width:180px"><b>{_esc(wer)}</b>{patetxt}</span>{tchip}</div>')
        offene_card = f'<div class="card">{mrows}</div>'
    else:
        offene_card = ('<div class="card"><p class="hint" style="margin:0">Keine offenen Maßnahmen. '
                       'In der Übersicht einen Ampelpunkt anklicken, um eine Schulung zu planen.</p></div>')

    # ── Lernbedarf je Person: Abdeckung + offene Punkte (rot/gelb) ──
    if aktive and aufgaben:
        kopf = ('<tr><th style="text-align:left">Person</th><th>grün</th><th>gelb</th>'
                '<th>rot</th><th>k.&nbsp;A.</th><th>abgedeckt</th></tr>')
        zeilen, details = '', ''
        for p in aktive:
            pid = str(p['id'])
            c = {'gruen': 0, 'gelb': 0, 'rot': 0, '': 0}
            offen = []
            for a in aufgaben:
                s = (bew.get(str(a['id'])) or {}).get(pid, '')
                c[s if s in AMPEL else ''] += 1
                if s in ('rot', 'gelb'):
                    offen.append((a.get('bereich') or '', a.get('name') or '', s))
            pct = round(c['gruen'] * 100 / n) if n else 0
            farbe = '#16a34a' if pct >= 80 else ('#eab308' if pct >= 50 else '#dc2626')
            zeilen += (f'<tr><td style="text-align:left"><b>{_esc(p["name"])}</b></td>'
                       f'<td>{c["gruen"]}</td><td>{c["gelb"]}</td><td>{c["rot"]}</td>'
                       f'<td>{c[""]}</td>'
                       f'<td style="color:{farbe};font-weight:700">{pct}%</td></tr>')
            if offen:
                items = ''.join(
                    '<div class="arow tp" style="padding:5px 0">'
                    f'<div class="an">{_dot(s)}&nbsp; '
                    f'<span style="color:var(--muted);font-size:11px">{_esc(b)}</span> {_esc(nm)}</div></div>'
                    for b, nm, s in offen)
                details += (
                    '<details style="margin:6px 0;border-top:1px solid var(--line);padding-top:6px">'
                    f'<summary style="cursor:pointer;font-weight:600">{_esc(p["name"])} &ndash; '
                    f'{len(offen)} offene Punkte (rot/gelb)</summary>'
                    f'<div style="padding:4px 0 8px">{items}</div></details>')
        lern = (f'<div class="card"><div class="tscroll"><table class="mtx" style="max-width:680px">'
                f'<thead>{kopf}</thead><tbody>{zeilen}</tbody></table></div>{details}</div>')
    else:
        lern = ('<div class="card"><p class="hint" style="margin:0">Noch keine Bewertungen – '
                'sobald das Team bewertet, erscheint hier der Lernbedarf.</p></div>')

    # ── Änderungsverlauf (Timeline, neueste zuerst) ──
    filter_btns = ''.join(
        f'<a class="btn secondary btn-sm" href="/verlauf?f={k}" '
        f'style="{"background:var(--accent);color:#fff;border-color:transparent" if f == k else ""}">{lbl}</a>'
        for k, lbl in (('', 'Alle'), ('gelernt', 'Gelernt ▲'), ('rueck', 'Rückschritte ▼')))
    rows = ''
    gezeigt = 0
    for e in reversed(hist):
        alt, neu = e.get('alt', ''), e.get('neu', '')
        ra, rn = RANG.get(alt, 0), RANG.get(neu, 0)
        art = 'gelernt' if rn > ra else ('rueck' if rn < ra else 'geaendert')
        if f == 'gelernt' and art != 'gelernt':
            continue
        if f == 'rueck' and art != 'rueck':
            continue
        if gezeigt >= 300:
            break
        gezeigt += 1
        nm, ber = tinfo.get(str(e.get('tid')), ('gelöschte Aufgabe', ''))
        if art == 'gelernt':
            chip = '<span class="chip2" style="background:#e7f6ec;color:#0f7a3d">gelernt ▲</span>'
        elif art == 'rueck':
            chip = '<span class="chip2" style="background:#fdecec;color:#b42318">Rückschritt ▼</span>'
        else:
            chip = '<span class="chip2" style="background:#eef1f4;color:#6b7280">geändert</span>'
        rows += (
            '<div class="arow" style="gap:10px">'
            f'<span style="color:var(--muted);font-size:12px;min-width:104px">{_esc(e.get("ts", ""))}</span>'
            f'<span style="min-width:130px"><b>{_esc(pname.get(str(e.get("pid")), "?"))}</b></span>'
            f'<span style="flex:1;min-width:180px"><span style="color:var(--muted);font-size:11px">'
            f'{_esc(ber)}</span> {_esc(nm)}</span>'
            f'<span style="white-space:nowrap">{_dot(alt)} &rarr; {_dot(neu)} &nbsp;{chip}</span></div>')
    if not rows:
        rows = '<p class="hint" style="margin:0">Noch keine Änderungen erfasst.</p>'
    verlauf = (f'<div class="row" style="gap:6px;margin:0 0 10px">{filter_btns}</div>'
               f'<div class="card">{rows}</div>')

    return (_platte('Teamleitung: Lernbedarf und was sich verändert hat')
            + _subtabs('verlauf', True)
            + '<div class="sect"><h2>Wochen-Report</h2><div class="rule"></div></div>'
            + '<p class="hint" style="margin:0 0 8px">Bewegung der letzten 7 Tage und aktuelle Abdeckung '
            'auf einen Blick.</p>'
            + wochen_card
            + '<div class="sect"><h2>Vertretungs-Warnung</h2><div class="rule"></div></div>'
            + '<p class="hint" style="margin:0 0 8px">Aufgaben, die <b>weniger als 2 Personen</b> „grün" '
            'können – das echte Urlaubs-/Krankheitsrisiko. Kritisch zuerst, dann längste ohne Bewegung.</p>'
            + risiko_card
            + '<div class="sect"><h2>Offene Maßnahmen</h2><div class="rule"></div></div>'
            + '<p class="hint" style="margin:0 0 8px">Geplante Schulungen mit Termin und Pate. '
            'Anklicken öffnet die Zelle. Überfällige sind rot markiert.</p>'
            + offene_card
            + '<div class="sect"><h2>Lernbedarf je Person</h2><div class="rule"></div></div>'
            + '<p class="hint" style="margin:0 0 8px">Abdeckung = Anteil der Aufgaben auf „grün". '
            'Aufklappen zeigt die offenen Punkte (rot/gelb) je Person – die Schulungsliste.</p>'
            + lern
            + '<div class="sect"><h2>Änderungsverlauf</h2><div class="rule"></div></div>'
            + '<p class="hint" style="margin:0 0 8px">Jede Bewertungsänderung mit Zeitstempel. '
            '„Gelernt ▲" = jemand kann jetzt mehr (rot→gelb→grün), „Rückschritt ▼" = umgekehrt.</p>'
            + verlauf)


# ── Zelle: Notizen & Maßnahme (Aufgabe × Person, nur Admin) ──────────────────
def _zelle_html(tid, pid):
    a = next((x for x in _aufgaben() if str(x['id']) == str(tid)), None)
    p = next((x for x in _personen() if str(x['id']) == str(pid)), None)
    if not a or not p:
        return None
    s = (_bewertungen().get(str(tid)) or {}).get(str(pid), '')
    stlabel = AMPEL[s][2] if s in AMPEL else 'keine Angabe'
    ck = _ckey(tid, pid)
    notizen = _notizen().get(ck, [])
    m = _massnahmen().get(ck, {})
    aktive = _personen(nur_aktiv=True)
    tp = _esc(tid) + '/' + _esc(pid)

    kopf = (f'<div class="card"><div style="font-weight:700;color:var(--ink)">{_esc(a.get("bereich") or "")}</div>'
            f'<h2 style="margin:2px 0 6px">{_esc(a["name"])}</h2>'
            f'<p style="margin:0">Person: <b>{_esc(p["name"])}</b> &nbsp;·&nbsp; '
            f'Aktuell: {_dot(s)} {stlabel}</p></div>')

    pate_opts = '<option value="">— kein Pate —</option>' + ''.join(
        f'<option value="{px["id"]}"{" selected" if str(m.get("pate")) == str(px["id"]) else ""}>'
        f'{_esc(px["name"])}</option>' for px in aktive)
    status = m.get('status') or 'offen'
    inp = 'padding:8px 10px;border:1px solid #d5dde5;border-radius:7px;font:inherit'
    massn_card = (
        '<div class="sect"><h2>Maßnahme / Schulung</h2><div class="rule"></div></div>'
        f'<div class="card"><form method="post" action="/zelle/{tp}/massnahme" '
        'class="row" style="gap:10px;flex-wrap:wrap;align-items:flex-end">'
        f'<label class="fld" style="margin:0"><span>Termin</span><input type="date" name="termin" '
        f'value="{_esc(m.get("termin") or "")}" style="{inp}"></label>'
        f'<label class="fld" style="margin:0"><span>Pate (schult)</span><select name="pate" style="{inp}">'
        f'{pate_opts}</select></label>'
        f'<label class="fld" style="margin:0"><span>Status</span><select name="status" style="{inp}">'
        f'<option value="offen"{" selected" if status == "offen" else ""}>offen</option>'
        f'<option value="erledigt"{" selected" if status == "erledigt" else ""}>erledigt</option>'
        '</select></label>'
        '<button class="btn" type="submit">Speichern</button>'
        '<button class="btn secondary" type="submit" name="loeschen" value="1">entfernen</button>'
        '</form></div>')

    if notizen:
        nlist = ''.join(
            '<div class="arow" style="gap:10px">'
            f'<span style="color:var(--muted);font-size:12px;min-width:104px">{_esc(x.get("ts", ""))}</span>'
            f'<span style="flex:1">{_esc(x.get("text", ""))}</span></div>' for x in reversed(notizen))
    else:
        nlist = '<p class="hint" style="margin:0">Noch keine Notiz.</p>'
    notiz_card = (
        '<div class="sect"><h2>Notizen</h2><div class="rule"></div></div>'
        f'<div class="card"><form method="post" action="/zelle/{tp}/notiz" class="row" style="gap:8px">'
        '<input name="text" placeholder="z. B. am 15.8. mit Tina geübt" required '
        f'style="flex:1;min-width:220px;{inp}"><button class="btn" type="submit">+ Notiz</button></form>'
        f'<div style="margin-top:10px">{nlist}</div></div>')

    zh = [e for e in _historie() if str(e.get('tid')) == str(tid) and str(e.get('pid')) == str(pid)]
    if zh:
        hrows = ''.join(
            '<div class="arow" style="gap:10px">'
            f'<span style="color:var(--muted);font-size:12px;min-width:104px">{_esc(e.get("ts", ""))}</span>'
            f'<span>{_dot(e.get("alt", ""))} &rarr; {_dot(e.get("neu", ""))}</span></div>'
            for e in reversed(zh))
    else:
        hrows = '<p class="hint" style="margin:0">Noch keine Änderungen.</p>'
    hist_card = ('<div class="sect"><h2>Verlauf dieser Zelle</h2><div class="rule"></div></div>'
                 f'<div class="card">{hrows}</div>')

    return (_platte('Aufgabe × Person: Notizen & Maßnahme')
            + '<p style="margin:0 0 8px"><a href="/verlauf">&larr; Teamleitung</a> &nbsp;·&nbsp; '
            '<a href="/">Übersicht</a></p>'
            + kopf + massn_card + notiz_card + hist_card)


# ── Routen ───────────────────────────────────────────────────────────────────
@app.get('/', response_class=HTMLResponse)
def matrix(request: Request):
    return HTMLResponse(_seite(_matrix_html(_is_admin(request.state.user)), request.state.user))


@app.get('/bewerten', response_class=HTMLResponse)
def bewerten(request: Request, person: str = ''):
    return HTMLResponse(_seite(_bewerten_html(person, _is_admin(request.state.user)),
                               request.state.user))


@app.post('/bewerten')
async def bewerten_speichern(request: Request):
    form = await request.form()
    pid = str(form.get('person') or '')
    if not pid:
        return RedirectResponse('/bewerten', status_code=303)
    bew = _bewertungen()
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    events = []
    for a in _aufgaben():
        val = form.get(f't{a["id"]}')
        neu = val if val in AMPEL else ''
        tkey = str(a['id'])
        eintrag = bew.get(tkey) or {}
        alt = eintrag.get(pid, '')
        if alt != neu:
            events.append({'ts': now, 'pid': pid, 'tid': tkey, 'alt': alt, 'neu': neu})
        if neu:
            eintrag[pid] = neu
        else:
            eintrag.pop(pid, None)
        if eintrag:
            bew[tkey] = eintrag
        elif tkey in bew:
            del bew[tkey]
    _sichere(BEWERTUNG_PATH, bew)
    _historie_add(events)
    return RedirectResponse(f'/bewerten?person={pid}&ok=1', status_code=303)


@app.post('/bewerten/pauschal')
async def bewerten_pauschal(request: Request):
    form = await request.form()
    pid = str(form.get('person') or '')
    if not pid:
        return RedirectResponse('/bewerten', status_code=303)
    stufe = form.get('stufe') if form.get('stufe') in AMPEL else ''
    bew = _bewertungen()
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    events = []
    for a in _aufgaben():
        tkey = str(a['id'])
        eintrag = bew.get(tkey) or {}
        alt = eintrag.get(pid, '')
        if alt != stufe:
            events.append({'ts': now, 'pid': pid, 'tid': tkey, 'alt': alt, 'neu': stufe})
        if stufe:
            eintrag[pid] = stufe
        else:
            eintrag.pop(pid, None)
        if eintrag:
            bew[tkey] = eintrag
        elif tkey in bew:
            del bew[tkey]
    _sichere(BEWERTUNG_PATH, bew)
    _historie_add(events)
    return RedirectResponse(f'/bewerten?person={pid}&ok=1', status_code=303)


@app.get('/verwalten', response_class=HTMLResponse)
def verwalten(request: Request, neu: str = '', pers: str = '', gruen: str = '', fehler: str = ''):
    meldung = ''
    if fehler:
        meldung = ('<div class="card" style="border-left:4px solid #dc2626;margin-bottom:14px">'
                   '<p style="margin:0">Import fehlgeschlagen – bitte eine .xlsx- oder .csv-Datei wählen.</p></div>')
    elif neu or pers or gruen:
        teile = []
        if neu:
            teile.append(f'<b>{_esc(neu)}</b> neue Aufgaben')
        if pers and pers != '0':
            teile.append(f'<b>{_esc(pers)}</b> neue Personen')
        if gruen and gruen != '0':
            teile.append(f'<b>{_esc(gruen)}</b>&times; „grün" aus Assignee gesetzt')
        meldung = ('<div class="card" style="border-left:4px solid #16a34a;margin-bottom:14px">'
                   f'<p style="margin:0">Import übernommen: {", ".join(teile) or "nichts Neues"}.</p></div>')
    return HTMLResponse(_seite(_verwalten_html(meldung, _is_admin(request.state.user)),
                               request.state.user))


@app.get('/verlauf', response_class=HTMLResponse)
def verlauf(request: Request, f: str = ''):
    if not _is_admin(request.state.user):
        return RedirectResponse('/', status_code=303)
    return HTMLResponse(_seite(_verlauf_html(f), request.state.user))


@app.get('/zelle/{tid}/{pid}', response_class=HTMLResponse)
def zelle_ansicht(request: Request, tid: str, pid: str):
    if not _is_admin(request.state.user):
        return RedirectResponse('/', status_code=303)
    html = _zelle_html(tid, pid)
    if html is None:
        return RedirectResponse('/', status_code=303)
    return HTMLResponse(_seite(html, request.state.user))


@app.post('/zelle/{tid}/{pid}/notiz')
async def zelle_notiz(request: Request, tid: str, pid: str):
    if not _is_admin(request.state.user):
        return RedirectResponse('/', status_code=303)
    form = await request.form()
    text = (form.get('text') or '').strip()
    if text:
        d = _notizen()
        ck = _ckey(tid, pid)
        lst = d.get(ck) or []
        lst.append({'ts': datetime.now().strftime('%Y-%m-%d %H:%M'), 'text': text[:500]})
        d[ck] = lst
        _sichere(NOTIZEN_PATH, d)
    return RedirectResponse(f'/zelle/{tid}/{pid}', status_code=303)


@app.post('/zelle/{tid}/{pid}/massnahme')
async def zelle_massnahme(request: Request, tid: str, pid: str):
    if not _is_admin(request.state.user):
        return RedirectResponse('/', status_code=303)
    form = await request.form()
    d = _massnahmen()
    ck = _ckey(tid, pid)
    if form.get('loeschen') == '1':
        d.pop(ck, None)
    else:
        d[ck] = {'termin': (form.get('termin') or '').strip(),
                 'pate': (form.get('pate') or '').strip(),
                 'status': form.get('status') if form.get('status') in ('offen', 'erledigt') else 'offen',
                 'erstellt': datetime.now().strftime('%Y-%m-%d %H:%M')}
    _sichere(MASSNAHMEN_PATH, d)
    return RedirectResponse(f'/zelle/{tid}/{pid}', status_code=303)


@app.post('/aufgabe')
async def aufgabe_add(request: Request):
    form = await request.form()
    name = (form.get('name') or '').strip()
    if name:
        liste = _aufgaben()
        liste.append({'id': secrets.token_hex(5), 'name': name,
                      'bereich': (form.get('bereich') or '').strip()})
        _sichere(AUFGABEN_PATH, liste)
    return RedirectResponse('/verwalten', status_code=303)


@app.post('/aufgabe/{aid}/loeschen')
def aufgabe_del(request: Request, aid: str):
    _sichere(AUFGABEN_PATH, [a for a in _aufgaben() if str(a.get('id')) != str(aid)])
    bew = _bewertungen()
    if str(aid) in bew:
        del bew[str(aid)]
        _sichere(BEWERTUNG_PATH, bew)
    return RedirectResponse('/verwalten', status_code=303)


@app.post('/person')
async def person_add(request: Request):
    if not _is_admin(request.state.user):
        return RedirectResponse('/verwalten', status_code=303)
    form = await request.form()
    name = (form.get('name') or '').strip()
    if name:
        liste = _personen()
        liste.append({'id': secrets.token_hex(5), 'name': name, 'aktiv': True})
        _sichere(PERSONEN_PATH, liste)
    return RedirectResponse('/verwalten', status_code=303)


@app.post('/person/{pid}/aktiv')
def person_aktiv(request: Request, pid: str):
    if not _is_admin(request.state.user):
        return RedirectResponse('/verwalten', status_code=303)
    liste = _personen()
    for p in liste:
        if str(p.get('id')) == str(pid):
            p['aktiv'] = not p.get('aktiv', True)
    _sichere(PERSONEN_PATH, liste)
    return RedirectResponse('/verwalten', status_code=303)


def _person_move(request: Request, pid: str, richtung: str):
    if not _is_admin(request.state.user):
        return RedirectResponse('/verwalten', status_code=303)
    liste = _personen()
    idx = next((i for i, p in enumerate(liste) if str(p.get('id')) == str(pid)), None)
    if idx is not None:
        ziel = idx - 1 if richtung == 'hoch' else idx + 1
        if 0 <= ziel < len(liste):
            liste[idx], liste[ziel] = liste[ziel], liste[idx]
            _sichere(PERSONEN_PATH, liste)
    return RedirectResponse('/verwalten', status_code=303)


@app.post('/person/{pid}/hoch')
def person_hoch(request: Request, pid: str):
    return _person_move(request, pid, 'hoch')


@app.post('/person/{pid}/runter')
def person_runter(request: Request, pid: str):
    return _person_move(request, pid, 'runter')


@app.post('/person/{pid}/backup')
def person_backup(request: Request, pid: str):
    """Person als Backup markieren - sie zaehlt dann nicht in die Abdeckung."""
    if not _is_admin(getattr(request.state, 'user', None)):
        return RedirectResponse('/verwalten', status_code=303)
    liste = _personen()
    for p in liste:
        if str(p.get('id')) == str(pid):
            p['backup'] = not p.get('backup')
            break
    _sichere(PERSONEN_PATH, liste)
    return RedirectResponse('/verwalten', status_code=303)


@app.post('/person/{pid}/befristung')
async def person_befristung(request: Request, pid: str):
    """Grund und Stichtag - ab dem Tag danach zaehlt die Person nicht mehr."""
    if not _is_admin(getattr(request.state, 'user', None)):
        return RedirectResponse('/verwalten', status_code=303)
    form = await request.form()
    hinweis = ' '.join(str(form.get('hinweis') or '').split())[:60]
    bis = str(form.get('bis') or '').strip()[:10]
    liste = _personen()
    for p in liste:
        if str(p.get('id')) == str(pid):
            p['hinweis'] = hinweis
            p['bis'] = bis
            break
    _sichere(PERSONEN_PATH, liste)
    return RedirectResponse('/verwalten', status_code=303)


@app.post('/person/{pid}/loeschen')
def person_del(request: Request, pid: str):
    if not _is_admin(request.state.user):
        return RedirectResponse('/verwalten', status_code=303)
    _sichere(PERSONEN_PATH, [p for p in _personen() if str(p.get('id')) != str(pid)])
    bew = _bewertungen()
    geaendert = False
    for tkey in list(bew.keys()):
        if str(pid) in bew[tkey]:
            del bew[tkey][str(pid)]
            geaendert = True
            if not bew[tkey]:
                del bew[tkey]
    if geaendert:
        _sichere(BEWERTUNG_PATH, bew)
    return RedirectResponse('/verwalten', status_code=303)


def _zellen_lesen(dateiname, roh):
    """Excel/CSV robust in eine Zeilen-Liste (Liste von Zellen-Listen) lesen."""
    name = (dateiname or '').lower()
    if name.endswith('.csv'):
        text = roh.decode('utf-8-sig', errors='replace')
        trenner = ';' if text.count(';') >= text.count(',') else ','
        return [[c.strip() for c in zeile.split(trenner)]
                for zeile in text.splitlines() if zeile.strip()]
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(roh), read_only=True, data_only=True)
    ws = wb.active
    zeilen = []
    for row in ws.iter_rows(values_only=True):
        zeilen.append(['' if c is None else str(c).strip() for c in row])
    wb.close()
    return zeilen


@app.post('/import')
async def importieren(request: Request, datei: UploadFile = File(...)):
    if not _is_admin(request.state.user):
        return RedirectResponse('/verwalten', status_code=303)
    roh = await datei.read()
    try:
        zeilen = _zellen_lesen(datei.filename or '', roh)
    except Exception:  # noqa: BLE001
        return RedirectResponse('/verwalten?fehler=1', status_code=303)
    if not zeilen:
        return RedirectResponse('/verwalten', status_code=303)

    # Kopfzeile in den ersten Zeilen suchen (Excel hat oft eine Leerzeile davor).
    # Erkannt werden: Aufgabe (auch „Name"/„Tätigkeit"/„Task"), Bereich (auch
    # „Section"/„Column"/„Kategorie") und Person (auch „Assignee"/„Bearbeiter").
    i_auf = i_ber = i_pers = None
    kopf_idx = None
    for zi, zeile in enumerate(zeilen[:6]):
        hs = [c.lower() for c in zeile]
        auf = ber = pers = None
        for i, h in enumerate(hs):
            if pers is None and any(x in h for x in ('assignee', 'zugewiesen', 'mitarbeiter',
                                                     'verantwortlich', 'bearbeiter', 'zuständig',
                                                     'zustaendig', 'wer macht')):
                pers = i
            elif auf is None and any(x in h for x in ('aufgabe', 'tätigkeit', 'taetigkeit',
                                                      'task', 'bezeichnung', 'name')):
                auf = i
            elif ber is None and any(x in h for x in ('bereich', 'kategorie', 'gruppe',
                                                      'section', 'column', 'spalte')):
                ber = i
        if auf is not None:
            i_auf, i_ber, i_pers, kopf_idx = auf, ber, pers, zi
            break

    if i_auf is None:
        i_auf = 0
        datenzeilen = zeilen  # keine Kopfzeile erkannt → erste Spalte = Aufgabe
    else:
        datenzeilen = zeilen[kopf_idx + 1:]

    aufgaben = _aufgaben()
    personen = _personen()
    bew = _bewertungen()
    # Schlüssel = Bereich + Name: derselbe Name (z. B. ein Portal) kann in
    # mehreren Bereichen eine eigene Aufgabe sein.
    def _akey(ber, name):
        return f'{(ber or "").strip().lower()}||{(name or "").strip().lower()}'
    auf_by_name = {_akey(a.get('bereich'), a.get('name')): a for a in aufgaben}
    pers_by_name = {(p.get('name') or '').strip().lower(): p for p in personen}

    neu_auf = neu_pers = gesetzt = 0
    for z in datenzeilen:
        name = (z[i_auf].strip() if (i_auf is not None and i_auf < len(z)) else '')
        if not name:
            continue
        ber = (z[i_ber].strip() if (i_ber is not None and i_ber < len(z)) else '')
        key = _akey(ber, name)
        a = auf_by_name.get(key)
        if a is None:
            a = {'id': secrets.token_hex(5), 'name': name, 'bereich': ber}
            aufgaben.append(a)
            auf_by_name[key] = a
            neu_auf += 1

        # Assignee → als Person anlegen und diese Aufgabe direkt auf „grün"
        # (wer eine Aufgabe heute macht, kann sie). Jederzeit änderbar.
        ass = (z[i_pers].strip() if (i_pers is not None and i_pers < len(z)) else '')
        if ass:
            pkey = ass.lower()
            p = pers_by_name.get(pkey)
            if p is None:
                p = {'id': secrets.token_hex(5), 'name': ass, 'aktiv': True}
                personen.append(p)
                pers_by_name[pkey] = p
                neu_pers += 1
            eintrag = bew.get(str(a['id'])) or {}
            if eintrag.get(str(p['id'])) != 'gruen':
                eintrag[str(p['id'])] = 'gruen'
                bew[str(a['id'])] = eintrag
                gesetzt += 1

    _sichere(AUFGABEN_PATH, aufgaben)
    _sichere(PERSONEN_PATH, personen)
    _sichere(BEWERTUNG_PATH, bew)
    return RedirectResponse(
        f'/verwalten?neu={neu_auf}&pers={neu_pers}&gruen={gesetzt}', status_code=303)


@app.get('/export.xlsx')
def export_xlsx(request: Request):
    aufgaben = _aufgaben()
    personen = _personen(nur_aktiv=True)
    bew = _bewertungen()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    farben = {'gruen': '16A34A', 'gelb': 'EAB308', 'rot': 'DC2626'}
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ampelmatrix'
    kopf = ['Hauptaufgabe', 'Teilaufgabe'] + [p['name'] for p in personen] + ['grün', 'Status']
    ws.append(kopf)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='4C1D95')
        c.alignment = Alignment(vertical='center')
    for a in aufgaben:
        row = [a.get('bereich') or '', a.get('name') or '']
        g = 0
        for p in personen:
            s = (bew.get(str(a['id'])) or {}).get(str(p['id']), '')
            row.append(AMPEL[s][2] if s in AMPEL else '')
            # Backup zaehlt nicht in die Abdeckung - sonst sieht eine Aufgabe,
            # die nur die Vertretung kann, im Export abgedeckt aus.
            if s == 'gruen' and not _ist_backup(p) and p.get('aktiv', True):
                g += 1
        status = 'kritisch' if g == 0 else ('Risiko' if g == 1 else 'ok')
        row += [g, status]
        ws.append(row)
        r = ws.max_row
        for j, p in enumerate(personen):
            s = (bew.get(str(a['id'])) or {}).get(str(p['id']), '')
            if s in farben:
                ws.cell(r, 3 + j).fill = PatternFill('solid', fgColor=farben[s])
                ws.cell(r, 3 + j).font = Font(color='FFFFFF')
    widths = [20, 40] + [16] * len(personen) + [8, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = 'C2'

    ablage = DATA_DIR / 'export'
    ablage.mkdir(parents=True, exist_ok=True)
    pfad = ablage / f'ampelmatrix_{secrets.token_hex(4)}.xlsx'
    wb.save(pfad)
    wb.close()

    def aufraeumen():
        try:
            pfad.unlink(missing_ok=True)
        except OSError:
            pass

    name = f'Retailaufgaben_Ampelmatrix_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    return FileResponse(
        pfad, filename=name,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        background=BackgroundTask(aufraeumen))
